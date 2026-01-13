"""
Excel文件处理模块
处理Excel文件的读取、查找学生、更新成绩等功能
支持旧的.xls格式和新的.xlsx格式
"""
import os
from typing import Optional, Tuple, List
import xlrd
from xlrd import open_workbook
import xlwt
from xlutils.copy import copy as xlutils_copy
import openpyxl
from openpyxl import load_workbook
from pypinyin import lazy_pinyin
from config import EXCEL_COLUMNS, HEADER_ROWS


def levenshtein_distance(s1: str, s2: str) -> int:
    """
    计算两个字符串的编辑距离（Levenshtein Distance）
    """
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)

    if len(s2) == 0:
        return len(s1)

    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            # j+1 instead of j since previous_row and current_row are one character longer
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row

    return previous_row[-1]


def name_to_pinyin(name: str) -> str:
    """
    将中文姓名转换为拼音（小写，无音调）
    """
    pinyin_list = lazy_pinyin(name)
    return ''.join(pinyin_list).lower()


def fuzzy_match_name(search_name: str, candidate_name: str, max_distance: int = 1) -> bool:
    """
    使用拼音进行姓名模糊匹配
    允许最多 max_distance 个字符的差异
    """
    search_pinyin = name_to_pinyin(search_name)
    candidate_pinyin = name_to_pinyin(candidate_name)

    distance = levenshtein_distance(search_pinyin, candidate_pinyin)
    return distance <= max_distance


class ExcelHandler:
    def __init__(self):
        self.file_path: Optional[str] = None
        self.is_xls: bool = False  # True for .xls, False for .xlsx

        # For .xls files
        self.xls_book: Optional[xlrd.Book] = None
        self.xls_sheet: Optional[xlrd.sheet.Sheet] = None
        self.xls_workbook: Optional[xlwt.Workbook] = None  # For writing

        # For .xlsx files
        self.xlsx_wb: Optional[openpyxl.Workbook] = None
        self.xlsx_ws: Optional[openpyxl.worksheet.worksheet.Worksheet] = None

        self.header_rows = HEADER_ROWS  # 从config读取表头行数

    def load_excel(self, file_path: str) -> bool:
        """
        加载Excel文件
        返回True表示成功，False表示失败
        """
        try:
            if not os.path.exists(file_path):
                print(f"文件不存在: {file_path}")
                return False

            if not os.access(file_path, os.R_OK):
                print(f"文件无法读取: {file_path}")
                return False

            self.file_path = file_path

            # 判断文件格式
            if file_path.lower().endswith('.xls'):
                self.is_xls = True
                return self._load_xls()
            elif file_path.lower().endswith('.xlsx'):
                self.is_xls = False
                return self._load_xlsx()
            else:
                print("不支持的文件格式，请使用.xls或.xlsx文件")
                return False

        except PermissionError:
            print("文件被其他程序占用，无法打开")
            return False
        except Exception as e:
            print(f"加载Excel文件失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _load_xls(self) -> bool:
        """加载.xls文件"""
        try:
            # 使用xlrd打开.xls文件
            self.xls_book = xlrd.open_workbook(self.file_path, formatting_info=True)
            self.xls_sheet = self.xls_book.sheet_by_index(0)

            if self.xls_sheet.nrows <= self.header_rows:
                print(f"Excel文件行数不足，至少需要{self.header_rows + 1}行（包含表头）")
                return False

            print(f"成功加载.xls文件，共{self.xls_sheet.nrows}行，{self.xls_sheet.ncols}列")
            return True
        except Exception as e:
            print(f"无法打开.xls文件: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _load_xlsx(self) -> bool:
        """加载.xlsx文件"""
        try:
            self.xlsx_wb = load_workbook(self.file_path)
            self.xlsx_ws = self.xlsx_wb.active

            if self.xlsx_ws.max_row <= self.header_rows:
                print(f"Excel文件行数不足，至少需要{self.header_rows + 1}行（包含表头）")
                return False

            print(f"成功加载.xlsx文件，共{self.xlsx_ws.max_row}行")
            return True
        except Exception as e:
            print(f"无法打开.xlsx文件: {e}")
            print("提示: 请确保文件未被其他程序打开")
            import traceback
            traceback.print_exc()
            return False

    def find_student_by_sequence(self, sequence: int) -> Optional[int]:
        """
        根据序号查找学生，返回Excel中的行号（1-based，包含表头）
        序号是从1开始的数据行序号（跳过表头）
        例如：序号1 = Excel第3行（跳过2行表头）
        """
        try:
            if sequence < 1:
                print(f"序号必须大于0: {sequence}")
                return None

            # 计算Excel行号：序号 + 表头行数
            row = sequence + self.header_rows

            # 验证行号是否有效
            if self.is_xls:
                if self.xls_sheet is None or row > self.xls_sheet.nrows:
                    print(f"序号超出范围: {sequence}（总共{self.get_total_students()}个学生）")
                    return None
            else:
                if self.xlsx_ws is None or row > self.xlsx_ws.max_row:
                    print(f"序号超出范围: {sequence}（总共{self.get_total_students()}个学生）")
                    return None

            return row

        except Exception as e:
            print(f"查找序号失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def find_student_by_name(self, name: str) -> Optional[int]:
        """
        根据姓名查找学生，返回Excel中的行号（1-based，包含表头）
        支持拼音模糊匹配：
        1. 先尝试精确匹配
        2. 如果失败，使用拼音模糊匹配（允许1个字符差异）
        """
        try:
            name = str(name).strip()
            if not name:
                return None

            # 第一步：精确匹配
            if self.is_xls:
                result = self._find_in_xls(EXCEL_COLUMNS['name'], name)
            else:
                result = self._find_in_xlsx(EXCEL_COLUMNS['name'], name)

            if result is not None:
                return result

            # 第二步：拼音模糊匹配
            print(f"精确匹配失败，尝试拼音模糊匹配: {name}")
            if self.is_xls:
                return self._fuzzy_find_in_xls(name)
            else:
                return self._fuzzy_find_in_xlsx(name)

        except Exception as e:
            print(f"查找姓名失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _find_in_xls(self, col_idx: int, search_value: str) -> Optional[int]:
        """在.xls文件中查找"""
        if self.xls_sheet is None:
            return None

        # 从数据行开始查找（跳过表头）
        for row_idx in range(self.header_rows, self.xls_sheet.nrows):
            cell_value = str(self.xls_sheet.cell_value(row_idx, col_idx)).strip()

            # 精确匹配
            if cell_value == search_value:
                return row_idx + 1  # 返回1-based行号

            # 尝试数字匹配（对于学号）
            if col_idx == EXCEL_COLUMNS['student_id']:
                # 提取数字部分
                cell_num = ''.join(filter(str.isdigit, cell_value))
                search_num = ''.join(filter(str.isdigit, search_value))
                if cell_num and search_num and cell_num == search_num:
                    return row_idx + 1

        # 如果精确匹配失败，尝试模糊匹配（包含关系）
        for row_idx in range(self.header_rows, self.xls_sheet.nrows):
            cell_value = str(self.xls_sheet.cell_value(row_idx, col_idx)).strip()
            if search_value in cell_value or cell_value in search_value:
                return row_idx + 1

        return None

    def _find_in_xlsx(self, col_idx: int, search_value: str) -> Optional[int]:
        """在.xlsx文件中查找"""
        if self.xlsx_ws is None:
            return None

        # 从数据行开始查找（跳过表头）
        for row_idx in range(self.header_rows + 1, self.xlsx_ws.max_row + 1):
            cell_value = str(self.xlsx_ws.cell(row=row_idx, column=col_idx + 1).value or '').strip()

            # 精确匹配
            if cell_value == search_value:
                return row_idx

            # 尝试数字匹配（对于学号）
            if col_idx == EXCEL_COLUMNS['student_id']:
                cell_num = ''.join(filter(str.isdigit, cell_value))
                search_num = ''.join(filter(str.isdigit, search_value))
                if cell_num and search_num and cell_num == search_num:
                    return row_idx

        # 如果精确匹配失败，尝试模糊匹配
        for row_idx in range(self.header_rows + 1, self.xlsx_ws.max_row + 1):
            cell_value = str(self.xlsx_ws.cell(row=row_idx, column=col_idx + 1).value or '').strip()
            if search_value in cell_value or cell_value in search_value:
                return row_idx

        return None

    def _fuzzy_find_in_xls(self, search_name: str) -> Optional[int]:
        """
        在.xls文件中使用拼音进行模糊匹配
        允许1个字符的编辑距离
        """
        if self.xls_sheet is None:
            return None

        name_col = EXCEL_COLUMNS['name']
        best_match_row = None
        best_distance = float('inf')

        # 遍历所有学生姓名
        for row_idx in range(self.header_rows, self.xls_sheet.nrows):
            cell_value = str(self.xls_sheet.cell_value(row_idx, name_col)).strip()

            if fuzzy_match_name(search_name, cell_value, max_distance=1):
                # 找到匹配的，计算距离
                search_pinyin = name_to_pinyin(search_name)
                candidate_pinyin = name_to_pinyin(cell_value)
                distance = levenshtein_distance(search_pinyin, candidate_pinyin)

                # 保留最佳匹配
                if distance < best_distance:
                    best_distance = distance
                    best_match_row = row_idx + 1  # 返回1-based行号

        if best_match_row:
            print(f"拼音模糊匹配成功: {search_name} -> 行{best_match_row} (编辑距离: {best_distance})")

        return best_match_row

    def _fuzzy_find_in_xlsx(self, search_name: str) -> Optional[int]:
        """
        在.xlsx文件中使用拼音进行模糊匹配
        允许1个字符的编辑距离
        """
        if self.xlsx_ws is None:
            return None

        name_col = EXCEL_COLUMNS['name']
        best_match_row = None
        best_distance = float('inf')

        # 遍历所有学生姓名
        for row_idx in range(self.header_rows + 1, self.xlsx_ws.max_row + 1):
            cell_value = str(self.xlsx_ws.cell(row=row_idx, column=name_col + 1).value or '').strip()

            if fuzzy_match_name(search_name, cell_value, max_distance=1):
                # 找到匹配的，计算距离
                search_pinyin = name_to_pinyin(search_name)
                candidate_pinyin = name_to_pinyin(cell_value)
                distance = levenshtein_distance(search_pinyin, candidate_pinyin)

                # 保留最佳匹配
                if distance < best_distance:
                    best_distance = distance
                    best_match_row = row_idx

        if best_match_row:
            print(f"拼音模糊匹配成功: {search_name} -> 行{best_match_row} (编辑距离: {best_distance})")

        return best_match_row

    def get_student_info(self, row: int) -> Optional[dict]:
        """
        获取指定行的学生信息
        row: Excel行号（1-based）
        """
        try:
            if self.is_xls:
                if self.xls_sheet is None or row < 1 or row > self.xls_sheet.nrows:
                    return None
                student_id = str(self.xls_sheet.cell_value(row - 1, EXCEL_COLUMNS['student_id']))
                name = str(self.xls_sheet.cell_value(row - 1, EXCEL_COLUMNS['name']))
            else:
                if self.xlsx_ws is None or row < 1 or row > self.xlsx_ws.max_row:
                    return None
                student_id = str(self.xlsx_ws.cell(row=row, column=EXCEL_COLUMNS['student_id'] + 1).value or '')
                name = str(self.xlsx_ws.cell(row=row, column=EXCEL_COLUMNS['name'] + 1).value or '')

            return {
                'row': row,
                'student_id': student_id,
                'name': name
            }
        except Exception as e:
            print(f"获取学生信息失败: {e}")
            return None

    def get_score(self, row: int, column_type: str) -> Optional[float]:
        """
        获取指定行的成绩
        row: Excel行号（1-based）
        column_type: 'regular_score' 或 'final_score'
        返回: 分数（可能是None）
        """
        try:
            # 确定列
            if column_type == 'regular_score':
                col_idx = EXCEL_COLUMNS['regular_score']
            elif column_type == 'final_score':
                col_idx = EXCEL_COLUMNS['final_score']
            else:
                return None

            if self.is_xls:
                if self.xls_sheet is None or row < 1 or row > self.xls_sheet.nrows:
                    return None
                cell_value = self.xls_sheet.cell_value(row - 1, col_idx)
            else:
                if self.xlsx_ws is None or row < 1 or row > self.xlsx_ws.max_row:
                    return None
                cell_value = self.xlsx_ws.cell(row=row, column=col_idx + 1).value

            # 尝试转换为浮点数
            if cell_value is None or cell_value == '':
                return None
            return float(cell_value)

        except Exception as e:
            print(f"获取成绩失败: {e}")
            return None

    def update_score(self, row: int, column_type: str, score: float) -> bool:
        """
        更新指定行的成绩
        row: Excel行号（1-based）
        column_type: 'regular_score' 或 'final_score'
        score: 分数（0-100）
        """
        try:
            # 验证分数范围
            score = float(score)
            if score < 0 or score > 100:
                print(f"成绩超出范围: {score}")
                return False

            # 确定列
            if column_type == 'regular_score':
                col_idx = EXCEL_COLUMNS['regular_score']
            elif column_type == 'final_score':
                col_idx = EXCEL_COLUMNS['final_score']
            else:
                print(f"未知的列类型: {column_type}")
                return False

            if self.is_xls:
                return self._update_xls_score(row, col_idx, score)
            else:
                return self._update_xlsx_score(row, col_idx, score)

        except Exception as e:
            print(f"更新成绩失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _update_xls_score(self, row: int, col_idx: int, score: float) -> bool:
        """更新.xls文件中的成绩"""
        if self.xls_sheet is None:
            return False

        # 将row从1-based转为0-based
        row_idx = row - 1

        if row_idx < self.header_rows or row_idx >= self.xls_sheet.nrows:
            print(f"行号超出范围: {row}")
            return False

        # 记录要更新的单元格
        # 注意：这里只是记录，实际写入在save_excel时进行
        if not hasattr(self, '_xls_updates'):
            self._xls_updates = {}

        self._xls_updates[(row_idx, col_idx)] = score
        print(f"已记录更新: 行{row}, 列{col_idx}, 分数{score}")
        return True

    def _update_xlsx_score(self, row: int, col_idx: int, score: float) -> bool:
        """更新.xlsx文件中的成绩"""
        if self.xlsx_ws is None:
            return False

        if row <= self.header_rows or row > self.xlsx_ws.max_row:
            print(f"行号超出范围: {row}")
            return False

        # openpyxl使用1-based索引
        self.xlsx_ws.cell(row=row, column=col_idx + 1, value=score)
        return True

    def save_excel(self) -> bool:
        """
        保存Excel文件
        """
        try:
            if self.file_path is None:
                print("未加载Excel文件")
                return False

            # 检查文件是否可写
            if not os.access(self.file_path, os.W_OK):
                print(f"文件不可写: {self.file_path}")
                return False

            if self.is_xls:
                return self._save_xls()
            else:
                return self._save_xlsx()

        except PermissionError:
            print("文件被其他程序打开，无法保存")
            print("提示: 请关闭Excel文件后重试")
            return False
        except Exception as e:
            print(f"保存Excel文件失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _save_xls(self) -> bool:
        """保存.xls文件"""
        try:
            if not hasattr(self, '_xls_updates') or not self._xls_updates:
                print("没有需要保存的更改")
                return True

            # 使用xlutils复制原始工作簿
            write_book = xlutils_copy(self.xls_book)
            write_sheet = write_book.get_sheet(0)

            # 应用所有更新
            for (row_idx, col_idx), value in self._xls_updates.items():
                write_sheet.write(row_idx, col_idx, value)
                print(f"写入: 行{row_idx + 1}, 列{col_idx + 1}, 值{value}")

            # 保存文件
            write_book.save(self.file_path)
            print(f"成功保存.xls文件: {self.file_path}")

            # 清空更新记录
            self._xls_updates = {}

            # 重新加载文件以保持同步
            return self._load_xls()

        except Exception as e:
            print(f"保存.xls文件失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _save_xlsx(self) -> bool:
        """保存.xlsx文件"""
        try:
            if self.xlsx_wb is None:
                print("未加载.xlsx文件")
                return False

            self.xlsx_wb.save(self.file_path)
            print(f"成功保存.xlsx文件: {self.file_path}")
            return True

        except Exception as e:
            print(f"保存.xlsx文件失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def get_total_students(self) -> int:
        """
        获取学生总数
        """
        try:
            if self.is_xls:
                if self.xls_sheet is None:
                    return 0
                # 总行数 - 表头行数
                return max(0, self.xls_sheet.nrows - self.header_rows)
            else:
                if self.xlsx_ws is None:
                    return 0
                return max(0, self.xlsx_ws.max_row - self.header_rows)
        except Exception as e:
            print(f"获取学生总数失败: {e}")
            return 0
